
import pandas as pd
import openpyxl

# Ruta al archivo Excel
excel_path = '/home/ubuntu/calculadora_presupuesto_viajes/app/data/Dinamica_Presupuesto_Viajero1234566.xlsx'

# Abrir el archivo para examinar las hojas
workbook = openpyxl.load_workbook(excel_path)
print("Hojas disponibles:", workbook.sheetnames)

# Examinar cada hoja
for sheet_name in workbook.sheetnames[:6]:  # Las primeras 6 hojas
    print(f"\n=== HOJA: {sheet_name} ===")
    
    # Leer con pandas
    df = pd.read_excel(excel_path, sheet_name=sheet_name)
    print(f"Dimensiones: {df.shape}")
    print("Columnas:", df.columns.tolist())
    
    if not df.empty:
        print("Primeras 3 filas:")
        print(df.head(3))
        
        # Mostrar datos de las columnas clave según la especificación
        key_cols = ['A', 'B', 'C', 'D', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
        print("\nColumnas por posición (primeras 3 filas):")
        for i, col in enumerate(df.columns[:21]):  # Primeras 21 columnas (A-U)
            if i < len(key_cols):
                print(f"Columna {key_cols[i]} ({i+1}): {col} -> {df.iloc[:3, i].tolist()}")
    print("-" * 50)

# Cerrar el archivo
workbook.close()
